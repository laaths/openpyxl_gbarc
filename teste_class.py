from openpyxl import load_workbook, Workbook

# Class para carregamento do arquivo de leitura
class loadArq:
    def __init__(self, planilha, aba_sheet):
        self.plan = planilha
        self.sheet = aba_sheet
        self.alfab = '0ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    # Carregar a planilha
    def loadPlan(self):
        return load_workbook(self.plan)

    # Carregar a Aba da Planilha
    def loadSheet(self):
        return self.loadPlan()[self.sheet]

    # Ativar edição da planilha
    def activePlan(self):
        return self.loadPlan().active

    # Salvar planilha carregada
    def savePlan(self):
        return self.loadPlan().save(self.plan)

    # Criar lista com itens da linha
    def listCols(self, proc, load):
        cols = []
        for cl in range(load.max_column ):
            for ln in range( load.max_row ):
                position = load.cell(row=ln+1, column=cl+1).value
                if position == proc:
                    for cl2 in range(load.max_column ):
                        cols.append(load.cell(row=ln+1, column=cl2+1).value)
                    return cols
                else:
                    pass
        return False

    # Criar lista com Colunas do painel
    def listColsPanel(self, load):
        cols = []
        for cl in range(load.max_column):
            cols.append(load.cell(row=1, column=cl+1).value)
        return cols


    # Criar lista com itens da coluna
    def listRows(self, proc, load):
        rows = []
        for ln in range(load.max_row):
            pass

    # Inserir dados na planilha
    def insertDadosRow(self, dest_panel, orig_panel, ws, orig_dados):
        dados = []
        #for plan in range(ws.max_row):
        for cl in range(len(dest_panel)):
            if dest_panel[cl] in orig_panel:
                #print('Existe na planilha')
                dados.append(orig_dados[orig_panel.index(dest_panel[cl])])
            elif dest_panel[cl] not in orig_panel:
                dados.append(None)
            else:
                pass
        return dados

# Classe para salvar o arquivo / provavelmente não utilizado
class saveArq:
    def __init__(self, planilha, aba_sheet):
        self.plan = planilha
        self.sheet = aba_sheet