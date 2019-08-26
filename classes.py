from openpyxl import load_workbook, Workbook

class cadastrar:
    def __init__(self, nome, serie):
        self.name = nome
        self.nserie = serie
        self.datenow = self.hm_gen()

    def hm_gen(self): #RETORNA DATA/HORA ATUAL
        from datetime import datetime
        data_comp_atual = datetime.now()
        dh_atual = data_comp_atual.strftime( "%D" ) #( "%X" )
        return str( dh_atual )

    def getNome(self):
        return self.name

    def getNserie(self):
        return self.nserie

    def getDatenow(self):
        return self.datenow

    def setNome(self, nome):
        self.name = nome

    def setNserie(self, serie):
        self.nserie = serie

    def setDatenow(self, date):
        self.datenow = date

    def cadastramento(self, aba, ws):
        x = 2
        while True:
            cel = str(x)
            if aba['A' + cel].value == None:
                ws['A' + cel] = self.name
                ws['B' + cel] = self.nserie
                ws['C' + cel] = self.hm_gen()
                break
            else:
                x += 1
        return

    def procValorCel(self, aba, vproc, col):
        x = 2
        while True:
            cel = str( x )
            if aba[col + cel].value == vproc:
                print(cel, aba[col+cel].value)
            elif aba[col + cel].value == None:
                print("Acabou")
                break
            else:
                pass
            x += 1
        return

class arquivos:
    def __init__(self):
        self.tecnicos = None
        self.excel = None

    def rArq(self,arq_name, sheet_nome):
        try:
            wb = load_workbook(arq_name)# NOME DO ARQUIVO
            return wb
        except IOError:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_nome
            ws['A1'] = "NOME"
            ws['B1'] = "Nº SERIE"
            ws['C1'] = "DATA"
            wb.save(arq_name)
            return wb

            #sheet_ranges = wb['Retiradas']  # NOME DA ABA
            #ws = wb.active
