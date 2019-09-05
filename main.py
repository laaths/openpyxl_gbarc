from class import loadArq, saveArq
from openpyxl import load_workbook, Workbook
from msg import msgs

# Procura e retorna lista com os dados
def procPlan(planilha, aba, proc):
    plan = loadArq( planilha, aba )
    ws = plan.loadSheet()
    tt = plan.listCols( proc, ws )
    return tt

# Verificação do arquivo de texto
def verifArq():
    dados = [
        "config/config.xlsx",
        "config1",
        "PLANILHA DE ORIGEM DE DADOS",
        "CAMINHO DA PASTA",
        "NOME DA PLANILHA XLSX",
        "ABA DA PLANILHA",
        "PLANILHA DE DESTINO DE DADOS",
        "LISTA DE CIRCUITOS PARA COPIA A PARTIR DA LINHA 10"
    ]
    try:
        wb = load_workbook( dados[0] )  # NOME DO ARQUIVO
        return wb
    except FileNotFoundError:
        try:
            import os
            os.mkdir( "config" )
        except FileExistsError:
            pass
        wb = Workbook()
        ws = wb.active
        ws.title = dados[1]
        ws['A1'] = dados[2]
        ws['A2'] = dados[3]
        ws['B2'] = dados[4]
        ws['C2'] = dados[5]
        ws['A5'] = dados[6]
        ws['A6'] = dados[3]
        ws['B6'] = dados[4]
        ws['C6'] = dados[5]
        ws['A9'] = dados[7]
        wb.save(dados[0])
        return wb

def insertPlan(orig_plan, orig_aba, dest_plan, dest_aba, proc):
    orig_dados = procPlan( orig_plan, orig_aba, proc )  # Retorna dados procurado
    dest_dados = procPlan( dest_plan, dest_aba, proc )  # Retorna dados procurado
    if orig_dados is False:
        listaux = ["001", "Não Encontrado"]
        return listaux
    elif dest_dados != False:
        listaux = ["002", "Existente na planilha" , dest_dados]
        return listaux
    else:
        plandest = loadArq( dest_plan, dest_aba )
        wbdest = plandest.loadPlan()
        planorig = loadArq( orig_plan, orig_aba )
        orig_panel = plandest.listColsPanel( planorig.loadSheet() )  # Retorna nomes da coluna do painel fixo
        dest_panel = plandest.listColsPanel( plandest.loadSheet() )  # Retorna nomes da coluna do painel fixo
        ws = wbdest.active
        dad_ext = plandest.insertDadosRow( dest_panel, orig_panel, ws,
                                           orig_dados )  # Retorna lista do conteudo ordenado pela painel de destino
        ws.append( dad_ext )  # Adiciona dados de uma lista na proxima linha vazia
        wbdest.save( dest_plan )  # Salva o conteudo na planilha
        return "000"

def procPlanPrint(orig_plan, orig_aba, dest_plan, dest_aba):
    proc = input("Digite UL/CCT/IP: ")
    orig_dados = procPlan( orig_plan, orig_aba, proc )  # Retorna dados procurado
    dest_dados = procPlan( dest_plan, dest_aba, proc )  # Retorna dados procurado
    if dest_dados != False:
        print( dest_dados[1] )  # Endereço
        print( dest_dados[3] )  # Cidade
        print( dest_dados[4] )  # Razão Social
        print( dest_dados[6] )  # CD / UL
        print( dest_dados[7] )  # Circuito
        print( dest_dados[8] )  # IP LAN
        print( dest_dados[12] ) # Loopback
        return
    elif orig_dados != False:
        print( orig_dados[9] )  # Endereço
        print( orig_dados[13] ) # Cidade
        print( orig_dados[6] )  # Razão Social
        print( orig_dados[2] )  # CD / UL
        print( orig_dados[20] ) # Circuito
        print( orig_dados[24] ) # IP LAN
        print( orig_dados[30] ) # Loopback
        input("\nPRESSIONE ENTER\n")
        return
    else:
        print("NÃO ENCONTRADO NAS PLANILHAS")
        return



def localizarDados():
    wb = verifArq()
    ws = wb.active
    if ws.cell( row=3, column=2 ).value is None:
        print( 'CADASTRE AS PLANILHAS A SEREM UTILIZADA!' )
    else:
        print( '\nPlanilha de ORIGEM:\n',
               ws.cell( row=3, column=2 ).value,
               '\nLocalizada em:\n',
               ws.cell( row=3, column=1 ).value )

        print( '\nPlanilha de DESTINO:\n',
               ws.cell( row=7, column=2 ).value,
               '\nLocalizada em:\n',
               ws.cell( row=7, column=1 ).value )

def loadConfPlanOrig():
    wb = verifArq()
    ws = wb.active
    orig = [ws.cell(row=3, column=1).value+ws.cell(row=3, column=2).value, ws.cell(row=3, column=3).value]
    return orig

def loadConfPlanDest():
    wb = verifArq()
    ws = wb.active
    orig = [ws.cell(row=7, column=1).value + ws.cell(row=7, column=2).value, ws.cell(row=7, column=3 ).value]
    return orig

def loadConfPlanProc():
    wb = verifArq()
    ws = wb.active
    cctlst = []
    for x in range(ws.max_row):
        if ws.cell(row=x+10, column=1).value is None:
            pass
        else:
            cctlst.append(ws.cell(row=x+10, column=1).value)
    return cctlst

def insertPlanList(orig_plan, orig_aba, dest_plan, dest_aba):
    proc = loadConfPlanProc()
    if proc == []:
        print("Preencha os circuitos na planilha")
        return
    else:
        for x in range(len(proc)):
            insert = insertPlan(orig_plan, orig_aba, dest_plan, dest_aba, proc[x])
            if insert[0] == "001":
                print( '-' * 25 )
                print(insert[1])
                print( '-' * 25 )
            elif insert[0] == "002":
                aux = insert[2]
                print(insert[1])
                for y in range(len(aux)):
                    if aux[y] != None:
                        print(aux[y])
                    else:
                        pass
                print('-'*25)
            else:
                pass
        print("CONCLUIDO")

# print(procPlan(dest[0], dest[1], proc))
# print(painel(orig[0], orig[1]))

orig = loadConfPlanOrig()
dest = loadConfPlanDest()
proc = 'PAE 0838723'

#print(insertPlan(orig[0], orig[1], dest[0], dest[1], proc))

# Menu de opções
def menuOpt():
    localizarDados()
    while True:
        msgs().getMenu()
        opt = input("OPÇÃO DESEJADA: ").upper()
        if opt == 'Q':
            break
        elif opt == 'A':  # COPIAR E COLAR LINHA DE UMA PLANILHA PRA OUTRA
            insertPlanList(orig[0], orig[1], dest[0], dest[1])
        elif opt == 'B':  # PROCURAR
            procPlanPrint(orig[0], orig[1], dest[0], dest[1])
        elif opt == 'C':  # INSERIR LOOPBACK
            pass
        elif opt == 'D':
            pass
        elif opt == 'E':
            localizarDados()
        elif opt == 'F':
            pass
        else:
            msgs().getMsg1()

menuOpt()