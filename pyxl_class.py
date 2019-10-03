from openpyxl import load_workbook, Workbook
import os

class excell():
    def __init__(self):
        self.origPlan = self.loadConfPlanOrig()
        self.origPlanPanel = self.listColsPanel(self.origPlan)
        self.destPlan = self.loadConfPlanDest()
        self.destPlanPanel = self.listColsPanel(self.destPlan)
        self.proc = self.loadConfPlanProc()
        self.ulProc = self.loadConfPlanUl()

    ### GETS ###

    # PLANILHA DE ORIGEM
    def getOrigPlan(self):
        return self.origPlan

    # PLANILHA DE DESTINO
    def getDestPlan(self):
        return self.destPlan

    # LISTA DE CIRCUITOS PARA PROCURAR
    def getProc(self):
        return self.proc

    # LISTA DE ULS PARA PROCURAR
    def getUlProc(self):
        return self.ulProc

    # Verificação do arquivo de texto
    def verifDirArq(self):
        dados = [
            "config/config.xlsx",
            "config1",
            "PLANILHA DE ORIGEM DE DADOS",
            "CAMINHO DA PASTA",
            "NOME DA PLANILHA XLSX",
            "ABA DA PLANILHA",
            "PLANILHA DE DESTINO DE DADOS",
            "LISTA DE CIRCUITOS PARA COPIA A PARTIR DA LINHA 10",
            "LISTA DE UL PARA PESQUISA A PARTIR DA LINHA 10"
        ]
        try:
            wb = load_workbook( dados[0] )  # NOME DO ARQUIVO
            ws = wb.active
            ws.title = dados[1]
            if ws.cell( row=3, column=1 ).value is None:
                print("\n CADASTRAR CAMINHO COMPLETO DA PLANILHA DE ORIGEM \n")
                ws['A3'] = "ex: C:/Users/Usuario/"
            if ws.cell( row=3, column=2 ).value is None:
                print("\n CADASTRAR NOME DA PLANILHA DE ORIGEM .xlsx \n")
                ws['B3'] = "ex: planilha.xlsx"
            if ws.cell( row=3, column=3 ).value is None:
                print("\n CADASTRAR ABA DE ORIGEM A SER UTILIZADA \n")
                ws['C3'] = "ex: Plan1"
            ####################################################################
            if ws.cell( row=7, column=1 ).value is None:
                print( "\n CADASTRAR CAMINHO COMPLETO DA PLANILHA DE DESTINO \n" )
                ws['A7'] = "ex: C:/Users/Usuario/"
            if ws.cell( row=7, column=2 ).value is None:
                print( "\n CADASTRAR NOME DA PLANILHA DE DESTINO .xlsx \n" )
                ws['B7'] = "ex: planilha.xlsx"
            if ws.cell( row=7, column=3 ).value is None:
                print( "\n CADASTRAR ABA DE DESTINO A SER UTILIZADA \n" )
                ws['C7'] = "ex: Plan1"
            if (ws.cell( row=3, column=1 ).value or
                ws.cell( row=3, column=2 ).value or
                ws.cell( row=3, column=3 ).value or
                ws.cell( row=7, column=1 ).value or
                ws.cell( row=7, column=2 ).value or
                ws.cell( row=7, column=3 ).value) is None:
                wb.save(dados[0])
            else:
                pass
            return wb
        except FileNotFoundError:
            try:
                os.mkdir("config")
            except FileExistsError:
                pass
            wb = Workbook()
            ws = wb.active
            ws.title = dados[1]
            ws['A1'] = dados[2]
            ws['A2'] = dados[3]
            ws['A3'] = "ex: C:/Users/Usuario/"
            ws['B2'] = dados[4]
            ws['B3'] = "ex: planilha.xlsx"
            ws['C2'] = dados[5]
            ws['C3'] = "ex: Plan1"
            ws['A5'] = dados[6]
            ws['A6'] = dados[3]
            ws['A7'] = "ex: C:/Users/Usuario/"
            ws['B6'] = dados[4]
            ws['B7'] = "ex: planilha.xlsx"
            ws['C6'] = dados[5]
            ws['C7'] = "ex: Plan1"
            ws['A9'] = dados[7]
            ws['B9'] = dados[8]
            wb.save( dados[0] )
            return wb

    # CRIA UMA LISTA COM AS COLUNAS DO PAINEL DA PLANILHA
    def listColsPanel(self, load):
        cols = []
        for cl in range( load.max_column ):
            cols.append( load.cell( row=1, column=cl + 1 ).value )
        return cols

    # CARREGA PLANILHA DE ORIGEM DOS DADOS
    def loadConfPlanOrig(self):
        wb = self.verifDirArq()
        ws = wb.active
        orig = load_workbook(ws.cell( row=3, column=1 ).value + ws.cell( row=3, column=2 ).value)[ws.cell( row=3, column=3 ).value]
        return orig

    # CARREGA PLANILHA DE DESTINO DOS DADOS
    def loadConfPlanDest(self):
        wb = self.verifDirArq()
        ws = wb.active
        orig = load_workbook(ws.cell( row=7, column=1 ).value + ws.cell( row=7, column=2 ).value)[ws.cell( row=7, column=3 ).value]
        return orig

    # LISTA DE CIRCUITOS
    def loadConfPlanProc(self):
        wb = self.verifDirArq()
        ws = wb.active
        cctlst = []
        for x in range( ws.max_row ):
            if ws.cell( row=x + 10, column=1 ).value is None:
                pass
            else:
                cctlst.append( ws.cell( row=x + 10, column=1 ).value )
        self.proc = cctlst
        procfrmt = self.formatCircuitos()
        self.proc = procfrmt
        return self.proc

    # LISTA DE ULS
    def loadConfPlanUl(self):
        wb = self.verifDirArq()
        ws = wb.active
        cctlst = []
        for x in range( ws.max_row ):
            if ws.cell( row=x + 10, column=2 ).value is None:
                pass
            else:
                cctlst.append( ws.cell( row=x + 10, column=2 ).value )
        return cctlst

    # FORMATA ENTRADA DE CIRCUITOS
    def formatCircuitos(self):
        proccompl = []
        for x in range(len(self.proc)):
            if self.proc[x][0:3].upper() == "PLT" or self.proc[x][0:3].upper() == "PAE" and (self.proc[x][3] == " " or self.proc[x][3] == "_") and self.proc[x][4:11].isdigit():
                proccompl.append(self.proc[x].replace("_", " ").upper())
            elif len(self.proc[x]) >= 12:
                for y in range(len(self.proc[x])):
                    try:
                        if (self.proc[x][y]+self.proc[x][y+1]+self.proc[x][y+2]).upper() == 'PAE' or (self.proc[x][y]+self.proc[x][y+1]+self.proc[x][y+2]).upper() == 'PLT':
                            proccompl.append(self.proc[x][y:y+11].replace( "_", " " ))
                        else:
                            pass
                    except IndexError:
                        pass
            elif len(self.proc[x]) == None:
                proccompl.append("VAZIO")
            else:
                pass
        return proccompl

    # LEITURA DO ARQUIVO DE CONFIGURAÇÃO
    def confProg(self):
        wb = self.verifDirArq()
        ws = wb.active
        print( '\n', ws.cell( row=1, column=1 ).value, '\n',
               ws.cell( row=3, column=2 ).value,
               '\n', ws.cell( row=2, column=1 ).value, '\n',
               ws.cell( row=3, column=1 ).value )

        print( '\n', ws.cell( row=5, column=1 ).value, '\n',
               ws.cell( row=7, column=2 ).value,
               '\n', ws.cell( row=6, column=1 ).value, '\n',
               ws.cell( row=7, column=1 ).value )
        print("\nCIRCUITOS A PROCURAR:")
        for x in range(len(self.proc)):
            print(self.proc[x])
        print("\nUL LOTERICAS A PROCURAR:")
        for x in range(len(self.ulProc)):
            print(self.ulProc[x])

    # CRIA UMA LISTA COM OS ITENS DA LINHA
    # COLUNA POR COLUNA
    def listCols(self, load, procItem):
        cols = []
        for cl in range( load.max_column ):
            for ln in range( load.max_row ):
                position = load.cell( row=ln + 1, column=cl + 1 ).value
                if position == procItem:
                    for cl2 in range( load.max_column ):
                        cols.append( load.cell( row=ln + 1, column=cl2 + 1 ).value )
                    return cols
                else:
                    pass
        return False

    # COMPARA E RETORNA O QUE TEM NA PLANILHA DE DESTINO
    # NÃO GRAVA O QUE NAO EXISTE NO PAINEL DE DESTINO
    def comparRows(self, orig_panel, dest_panel, orig_dados):
        dados = []
        for cl in range( len( dest_panel ) ):
            if dest_panel[cl] in orig_panel:
                dados.append( orig_dados[orig_panel.index( dest_panel[cl] )] )
            elif dest_panel[cl] not in orig_panel:
                dados.append( None )
            else:
                pass
        return dados

    # # RETORNA LISTA COM OS NUMEROS DE INDEX DA PLANILHA
    def indexColl(self, panel):
        listIndex = []
        for cl in range( len( panel ) ):
            listIndex.append( panel.index( panel[cl] ) )
        return listIndex

    # PROCURAR ITEM SIMPLES/UNICO
    def procItemSimple(self, procItem):
        orig_dados = self.listCols( self.origPlan, procItem )  # Retorna dados procurado
        orig_panel = self.origPlanPanel
        dest_dados = self.listCols( self.destPlan, procItem )  # Retorna dados procurado
        dest_panel = self.destPlanPanel

        if dest_dados != False:
            print( "ENCONTRADO", self.destPlan )
            indexItens = self.indexColl( dest_panel )
            for x in range( len( indexItens ) ):
                if dest_dados[indexItens[x]] != None:
                    print( dest_panel[x] + ":", dest_dados[indexItens[x]] )
                else:
                    pass
            return
        elif orig_dados != False:
            print( "ENCONTRADO", self.origPlan )
            indexItens = self.indexColl( orig_panel )
            for x in range( len( indexItens ) ):
                if orig_dados[indexItens[x]] != None:
                    print( orig_panel[x] + ":", orig_dados[indexItens[x]] )
                else:
                    pass
            return
        else:
            print( "\nNÃO ENCONTRADO NAS PLANILHAS" )
            return

    # INSERE OS DADOS PROCURADOS NA PLANILHAD E DESTINO
    def insertPlanDados(self):
        wdest = self.destPlan
        for qtd in range( len( self.proc ) ):
            orig_dados = self.listCols( self.origPlan, self.proc[qtd] )  # Retorna dados procurado
            dest_dados = self.listCols( self.destPlan, self.proc[qtd] )  # Retorna dados procurado
            if orig_dados is False:
                print( "NÃO ENCONTRADO\n" )
            elif dest_dados != False:
                print( "\nEXISTENTE NA PLANILHA" )
                self.procItemSimple( self.proc[qtd] )
            else:
                dadInsert = self.comparRows( self.origPlanPanel, self.destPlanPanel, orig_dados )
                ws.append( dadInsert )
        wdest.save( self.destPlan )
        return
